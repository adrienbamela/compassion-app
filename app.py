from flask import Flask, render_template, request, redirect, url_for, session, send_file, abort, jsonify
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from flask import jsonify 
import shutil

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'devsecret')

DATA_DIR = os.environ.get('DATA_DIR', '.')
EXCEL_FILENAME = os.environ.get('EXCEL_FILENAME', 'presences_questions.xlsx')
EXCEL_FILE = os.path.join(DATA_DIR, EXCEL_FILENAME)
os.makedirs(DATA_DIR, exist_ok=True)

ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')

# --- Initialiser le fichier Excel si n'existe pas ---
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Présences"
    ws1.append(["Timestamp","Nom","Prénom","Sexe","Département","Adresse","Téléphone","Email","Événement","Responsable"])
    ws2 = wb.create_sheet(title="Questions")
    ws2.append(["Timestamp","Nom","Question","Événement"])
    ws3 = wb.create_sheet(title="Ouvriers")
    ws3.append(["Timestamp","Nom","Prénom","Sexe","Adresse","Téléphone","Email","Départements"])
    ws4 = wb.create_sheet(title="Nouveaux")
    ws4.append(["Timestamp","Nom","Prénom","Sexe","Quartier","Téléphone","Événement"])
    wb.save(EXCEL_FILE)

# --- Compléter les colonnes Excel ---
def ensure_excel_columns():
    wb = load_workbook(EXCEL_FILE)
    sheet_columns = {
        "Présences": 10,
        "Questions": 4,
        "Ouvriers": 8,
        "Nouveaux": 7
    }
    changed = False
    for sheet, cols in sheet_columns.items():
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                row = list(row)
                if len(row) < cols:
                    row += [""] * (cols - len(row))
                    for col_index, value in enumerate(row, start=1):
                        ws.cell(row=i, column=col_index, value=value)
                    changed = True
    if changed:
        wb.save(EXCEL_FILE)

ensure_excel_columns()

# --- Page d'accueil ---
@app.route('/')
def index():
    user_type = session.get('user_type', None)
    return render_template('index.html', user_type=user_type)

# --- Routes login test ---
@app.route('/login/<role>')
def login(role):
    if role in ['admin', 'nouveau', 'fidele', 'president']:
        session['user_type'] = role
        return redirect(url_for('index'))
    return "Rôle inconnu"

@app.route('/logout')
def logout():
    session.pop('user_type', None)
    session.pop('admin', None)
    session.pop('president', None)
    session.pop('president_fullname', None)
    return redirect(url_for('index'))

# --- Nouveaux venus ---
@app.route('/nouveaux', methods=['GET','POST'])
def nouveaux():
    if request.method == 'POST':
        nom = request.form['nom']
        prenom = request.form['prenom']
        sexe = request.form['sexe']
        quartier = request.form.get('quartier','')
        telephone = request.form['telephone']
        evenement = request.form.get('event','Premier Culte')
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = load_workbook(EXCEL_FILE)
        ws = wb["Nouveaux"]
        ws.append([timestamp, nom, prenom, sexe, quartier, telephone, evenement])
        wb.save(EXCEL_FILE)
        return render_template('merci.html', message="Bienvenue ! Vos informations ont été enregistrées.")
    return render_template('nouveaux.html')

# --- Questions ---
@app.route('/questions', methods=['GET','POST'])
def questions():
    if request.method == 'POST':
        nom = request.form['nom']
        question = request.form['question']
        evenement = request.form.get('event','CD')
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = load_workbook(EXCEL_FILE)
        ws = wb["Questions"]
        ws.append([timestamp, nom, question, evenement])
        wb.save(EXCEL_FILE)
        return render_template('merci.html', message="Question envoyée !")
    return render_template('questions.html')

# --- Témoignages ---
@app.route('/temoignage', methods=['GET','POST'])
def temoignage():
    if request.method == 'POST':
        nom = request.form['nom']
        temoignage_text = request.form['temoignage']
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = load_workbook(EXCEL_FILE)
        if "Temoignages" not in wb.sheetnames:
            ws = wb.create_sheet("Temoignages")
            ws.append(["Timestamp","Nom","Temoignage"])
        else:
            ws = wb["Temoignages"]

        ws.append([timestamp, nom, temoignage_text])
        wb.save(EXCEL_FILE)

        return render_template('merci.html', message="Merci pour votre témoignage !")
    return render_template('temoignage.html')

# --- Inscription ouvriers ---
@app.route('/inscription_ouvrier', methods=['GET','POST'])
def inscription_ouvrier():
    if request.method == 'POST':
        nom = request.form['nom']
        prenom = request.form['prenom']
        sexe = request.form['sexe']
        adresse = request.form.get('adresse','')
        telephone = request.form['telephone']
        email = request.form.get('email','')
        departements = request.form.getlist('departements') or ["Non défini"]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = load_workbook(EXCEL_FILE)
        ws = wb["Ouvriers"]
        ws.append([timestamp, nom, prenom, sexe, adresse, telephone, email, ', '.join(departements)])
        wb.save(EXCEL_FILE)
        return render_template('merci.html', message=f"Ouvrier {nom} {prenom} enregistré avec succès !")
    return render_template('ouvriers_inscription.html')

# --- Checklist ouvriers ---
@app.route('/ouvriers', methods=['GET','POST'])
def ouvriers():
    departements = ["LA JEUNESSE","LA CHORALE","LE PROTOCOLE","LA SECURITE","HYGIENE ET SALUBRITE"]

    if request.method == 'POST':
        departement = request.form.get('departement','Non défini')
        checked = request.form.getlist('ouvrier')
        evenement = request.form.get('event','CD')
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = load_workbook(EXCEL_FILE)
        ws_ouv = wb["Ouvriers"]
        ws_pres = wb["Présences"]

        for item in checked:
            nom_item, prenom_item = item.split('|')
            info = [r for r in ws_ouv.iter_rows(min_row=2, values_only=True) if r[1]==nom_item and r[2]==prenom_item]
            if info:
                sexe = info[0][3] if len(info[0])>3 else ''
                adresse = info[0][4] if len(info[0])>4 else ''
                telephone = info[0][5] if len(info[0])>5 else ''
                email = info[0][6] if len(info[0])>6 else ''
                ws_pres.append([timestamp, nom_item, prenom_item, sexe, departement, adresse, telephone, email, evenement, ''])
        wb.save(EXCEL_FILE)
        return render_template('merci.html', message=f"{len(checked)} présences enregistrées !")

    return render_template('ouvriers.html', departements=departements)

# --- Route AJAX pour lister les ouvriers par département ---
@app.route('/ouvriers/liste/<departement>')
def ouvriers_liste(departement):
    wb = load_workbook(EXCEL_FILE)
    ws_ouv = wb["Ouvriers"]
    ouvriers_list = []
    for r in ws_ouv.iter_rows(min_row=2, values_only=True):
        nom = r[1] if len(r) > 1 else ""
        prenom = r[2] if len(r) > 2 else ""
        deps = r[7].split(', ') if len(r) > 7 else []
        if departement in deps:
            ouvriers_list.append({"nom": nom, "prenom": prenom})
    return jsonify(ouvriers_list)

# --- Admin login ---
@app.route('/admin', methods=['GET','POST'])
def admin():
    if request.method == 'POST':
        password = request.form['password']
        if password == ADMIN_PASSWORD:
            session['admin'] = True
            session['user_type'] = 'admin'
            return redirect(url_for('admin_dashboard'))
        else:
            return render_template('admin.html', error="Mot de passe incorrect")
    return render_template('admin.html')

# --- Admin dashboard ---

BACKUP_DIR = os.path.join(DATA_DIR, "backup")
os.makedirs(BACKUP_DIR, exist_ok=True)

@app.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('admin'):
        return redirect(url_for('admin'))
    wb = load_workbook(EXCEL_FILE)
    presences = [dict(zip(["Timestamp","Nom","Prénom","Sexe","Département","Adresse","Téléphone","Email","Événement","Responsable"], row)) 
                 for row in wb["Présences"].iter_rows(values_only=True) if row[0] != "Timestamp"]
    questions = [dict(zip(["Timestamp","Nom","Question","Événement"], row)) 
                 for row in wb["Questions"].iter_rows(values_only=True) if row[0] != "Timestamp"]
    if "Temoignages" in wb.sheetnames:
        temoignages = [dict(zip(
            ["Timestamp","Nom","Temoignage"], 
            row))
            for row in wb["Temoignages"].iter_rows(values_only=True) if row[0] != "Timestamp"]
    else:
        temoignages = []
    return render_template('dashboard_admin.html', presences=presences, questions=questions,temoignages=temoignages)

# --- Télécharger Excel ---
@app.route('/admin/download')
def admin_download():
    if not session.get('admin') and session.get('user_type') != 'president':
        abort(403)
    return send_file(EXCEL_FILE, as_attachment=True)

# --- Infos sur l'église ---
@app.route('/infos_eglise')
def infos_eglise():
    return render_template('infos_eglise.html')


# --- Lancement Flask ---
if __name__ == "__main__":
    app.run(debug=True)
