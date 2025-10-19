from openpyxl import load_workbook

# Chemin vers ton fichier Excel
EXCEL_FILE = 'presences_questions.xlsx'

# Charger le classeur
wb = load_workbook(EXCEL_FILE)

if "Présences" in wb.sheetnames:
    ws = wb["Présences"]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row = list(row)
        if len(row) < 10:
            # Complète la ligne avec des chaînes vides pour atteindre 10 colonnes
            row += [""] * (10 - len(row))
            # Mettre à jour la ligne dans la feuille
            for col_index, value in enumerate(row, start=1):
                ws.cell(row=i, column=col_index, value=value)
    wb.save(EXCEL_FILE)
    print("Toutes les lignes de 'Présences' ont été mises à 10 colonnes ✅")
else:
    print("La feuille 'Présences' n'existe pas ❌")
